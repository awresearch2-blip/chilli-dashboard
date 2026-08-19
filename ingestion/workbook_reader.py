"""Opens the mastersheet workbook safely and loads the declarative sheet config.

Handles the real-world failure modes of a workbook that lives in OneDrive and
gets edited by hand: transient file locks while Excel is still writing,
partial saves, and Excel's `~$...` lock/temp files. Never crashes the caller
on a bad read -- retries with backoff and raises a clear error only after
exhausting retries.
"""

import time
from pathlib import Path

import openpyxl
import yaml

from utils.logging_config import get_logger
from utils.paths import SHEETS_CONFIG_PATH, WORKBOOK_PATH

logger = get_logger("workbook_reader")

LOCK_FILE_PREFIX = "~$"


class WorkbookReadError(Exception):
    """Raised when the workbook cannot be read after all retries are exhausted."""


def is_lock_file(path: Path) -> bool:
    return path.name.startswith(LOCK_FILE_PREFIX)


def load_sheet_config(config_path: Path = SHEETS_CONFIG_PATH) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def open_workbook(
    path: Path = WORKBOOK_PATH,
    max_retries: int = 5,
    retry_delay_seconds: float = 2.0,
):
    """Open the workbook in both data_only and formula modes.

    Returns (wb_values, wb_formulas). wb_values has computed values for
    formula cells (what analytics should read); wb_formulas is used only to
    detect a stale/missing formula cache (a formula cell whose cached value
    is None because the workbook was never opened+saved in Excel since the
    formula was entered).
    """
    if is_lock_file(path):
        raise WorkbookReadError(f"Refusing to read Excel lock file: {path}")

    last_error = None
    for attempt in range(1, max_retries + 1):
        try:
            wb_values = openpyxl.load_workbook(path, data_only=True, read_only=False)
            wb_formulas = openpyxl.load_workbook(path, data_only=False, read_only=False)
            return wb_values, wb_formulas
        except (PermissionError, OSError, KeyError) as exc:
            # PermissionError: file locked by Excel. OSError/KeyError: mid-write /
            # corrupted zip while OneDrive/Excel is still flushing the save.
            last_error = exc
            logger.warning(
                "Workbook read attempt %d/%d failed (%s: %s) -- retrying in %.1fs",
                attempt, max_retries, type(exc).__name__, exc, retry_delay_seconds,
            )
            time.sleep(retry_delay_seconds)

    raise WorkbookReadError(
        f"Could not read workbook after {max_retries} attempts: {last_error}"
    ) from last_error


def discover_sheets(wb) -> list[str]:
    return list(wb.sheetnames)


def find_unconfigured_sheets(wb, sheet_config: dict) -> list[str]:
    configured = set(sheet_config.get("sheets", {}).keys())
    return [name for name in discover_sheets(wb) if name not in configured]


def detect_stale_formula_cache(ws_values, ws_formulas) -> int:
    """Count cells that are formulas but whose cached computed value is missing.

    This happens when a formula is entered/changed by something other than
    Excel itself (or the file is saved by a tool that doesn't recompute), so
    data_only=True reads None instead of a real number. Not fatal, but worth
    surfacing in the Data Quality Report.
    """
    stale_count = 0
    max_row = ws_formulas.max_row
    max_col = ws_formulas.max_column
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            f_cell = ws_formulas.cell(row=row, column=col)
            if f_cell.data_type == "f" and ws_values.cell(row=row, column=col).value is None:
                stale_count += 1
    return stale_count
