"""Reshapes each worksheet into a tidy long-format DataFrame per its declarative
spec in config/sheets.yaml.

This is a *structural* transformation only -- cell values are carried through
exactly as read (no type coercion, no token-to-NaN conversion, no dedup).
That happens in validation/cleaning. This module's job is purely to turn
hand-built merged-header pivot tables into queryable rows without guessing
at anything not stated in the config.
"""

import datetime as dt

import openpyxl.utils
import pandas as pd

from utils.logging_config import get_logger

logger = get_logger("sheet_parser")

MONTH_MAP = {
    name.lower(): i + 1
    for i, name in enumerate(
        ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    )
}


def _month_to_num(value):
    if value is None:
        return None
    key = str(value).strip().lower()[:3]
    return MONTH_MAP.get(key)


def _is_4digit_year(value) -> bool:
    try:
        return 1900 <= int(value) <= 2100
    except (TypeError, ValueError):
        return False


def _safe_date(year, month, day=1):
    try:
        return pd.Timestamp(year=int(year), month=int(month), day=day)
    except (TypeError, ValueError):
        return pd.NaT


def _id_column_name(spec: dict) -> str:
    id_spec = spec["id_column"]
    return id_spec.get("rename") or id_spec["role"]


def unmerge_row_values(ws, row_idx: int, max_col: int) -> list:
    """Read a header row with merged cells fully expanded (both directions)."""
    values = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row_idx <= merged_range.max_row:
            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                values[c - 1] = top_left
    return values


def parse_long(ws, spec: dict) -> pd.DataFrame:
    header_row = spec.get("header_row", 1)
    data_start_row = spec.get("data_start_row", header_row + 1)
    max_col = ws.max_column
    headers = unmerge_row_values(ws, header_row, max_col)
    id_idx = spec["id_column"]["index"]
    id_name = _id_column_name(spec)
    col_specs = {k.strip().lower(): v for k, v in spec.get("columns", {}).items()}

    records = []
    for row in range(data_start_row, ws.max_row + 1):
        row_id = ws.cell(row=row, column=id_idx).value
        record = {"_row": row, id_name: row_id}
        any_value = row_id is not None
        for col in range(1, max_col + 1):
            if col == id_idx:
                continue
            header_text = headers[col - 1]
            if header_text is None:
                continue
            info = col_specs.get(str(header_text).strip().lower())
            if info is None:
                continue
            value = ws.cell(row=row, column=col).value
            record[info["rename"]] = value
            if value is not None:
                any_value = True
        if not any_value:
            continue
        records.append(record)

    df = pd.DataFrame(records)
    if spec.get("trim_trailing_blank") and id_name in df.columns:
        df = df[df[id_name].notna()].reset_index(drop=True)
    if spec.get("sort") == "ascending" and id_name in df.columns:
        df = df.sort_values(id_name).reset_index(drop=True)
    return df


def parse_wide_pivot_year_month(ws, spec: dict) -> pd.DataFrame:
    id_idx = spec["id_column"]["index"]
    id_role = spec["id_column"]["role"]  # "year" or "month_name"
    header_row = spec.get("header_row", 1)
    data_start_row = spec.get("data_start_row", header_row + 1)
    max_col = spec.get("max_data_column", ws.max_column)
    headers = unmerge_row_values(ws, header_row, max_col)
    filters = spec.get("data_row_filter") or []
    if isinstance(filters, str):
        filters = [filters]
    value_name = spec.get("value_name", "value")

    records = []
    for row in range(data_start_row, ws.max_row + 1):
        id_value = ws.cell(row=row, column=id_idx).value
        if "stop_at_blank_id" in filters and id_value is None:
            # A second, independent table starts further down this sheet
            # (e.g. a unit-converted duplicate) -- stop rather than flatten it in.
            break
        if id_value is None:
            continue
        if "id_is_4digit_year" in filters and not _is_4digit_year(id_value):
            continue
        if "exclude_total_row" in filters and str(id_value).strip().lower() == "total":
            continue
        for col in range(1, max_col + 1):
            if col == id_idx:
                continue
            period_label = headers[col - 1]
            if period_label is None:
                continue
            value = ws.cell(row=row, column=col).value
            records.append(
                {"_row": row, id_role: id_value, "period_label": period_label, value_name: value}
            )

    df = pd.DataFrame(records)
    if df.empty:
        return df

    if id_role == "year":
        years, months = df["year"], df["period_label"].map(_month_to_num)
    else:
        years, months = df["period_label"], df[id_role].map(_month_to_num)
    df["date"] = [_safe_date(y, m) for y, m in zip(years, months)]
    return df


def parse_wide_pivot_year_cols(ws, spec: dict) -> pd.DataFrame:
    id_idx = spec["id_column"]["index"]
    id_name = _id_column_name(spec)
    header_row = spec.get("header_row", 1)
    data_start_row = spec.get("data_start_row", header_row + 1)
    max_col = ws.max_column
    headers = unmerge_row_values(ws, header_row, max_col)
    estimate_cols = {str(x).strip() for x in spec.get("estimate_columns", [])}
    value_name = spec.get("value_name", "value")

    records = []
    for row in range(data_start_row, ws.max_row + 1):
        id_value = ws.cell(row=row, column=id_idx).value
        if id_value is None:
            continue
        for col in range(1, max_col + 1):
            if col == id_idx:
                continue
            header_value = headers[col - 1]
            if header_value is None:
                continue
            value = ws.cell(row=row, column=col).value
            records.append(
                {
                    "_row": row,
                    id_name: str(id_value).strip(),
                    # Cast to str: years are ints except the trailing "(exp)"
                    # estimate column, which is text -- same mixed-type risk
                    # as the merged-header "group" field above.
                    "year": str(header_value).strip(),
                    value_name: value,
                    "is_estimate": str(header_value).strip() in estimate_cols,
                }
            )
    return pd.DataFrame(records)


def parse_wide_pivot_merged_header(ws, spec: dict) -> pd.DataFrame:
    id_idx = spec["id_column"]["index"]
    id_name = _id_column_name(spec)
    group_row_idx, metric_row_idx = spec["header_rows"]
    data_start_row = spec["data_start_row"]
    max_col = ws.max_column
    group_headers = unmerge_row_values(ws, group_row_idx, max_col)
    metric_headers = (
        unmerge_row_values(ws, metric_row_idx, max_col)
        if metric_row_idx != group_row_idx
        else group_headers
    )
    metric_roles = {k.strip().lower(): v for k, v in spec.get("metric_roles", {}).items()}
    column_overrides = spec.get("column_overrides", {})
    estimate_groups = {str(x).strip() for x in spec.get("estimate_groups", [])}
    group_role = spec.get("group_role", "group")

    col_plan = {}
    for col in range(1, max_col + 1):
        if col == id_idx:
            continue
        col_letter = openpyxl.utils.get_column_letter(col)
        if col_letter in column_overrides:
            ov = column_overrides[col_letter]
            col_plan[col] = {
                "name": ov["name"],
                "role": ov.get("role", "unknown"),
                "group": ov.get("group"),
                "is_estimate": False,
            }
            continue

        group_val = group_headers[col - 1]
        metric_val = metric_headers[col - 1]
        if group_val is None and metric_val is None:
            continue

        metric_key = str(metric_val).strip().lower() if metric_val is not None else None
        role_info = metric_roles.get(metric_key)
        if role_info is not None:
            final_name, role = role_info["rename"], role_info["role"]
        else:
            final_name = str(metric_val if metric_val is not None else group_val).strip()
            role = "unknown"

        # Cast to str: Excel stores some group headers as numbers (e.g. the
        # "334" variety) and others as text (e.g. "2026(exp)") in the same
        # header row -- an un-cast column would silently mix int/str types,
        # which breaks any downstream string-based filter (variety == "334")
        # unless something else happens to coerce it first. Never rely on that.
        group_val_str = str(group_val).strip() if group_val is not None else None
        is_estimate = group_val_str is not None and group_val_str in estimate_groups
        col_plan[col] = {"name": final_name, "role": role, "group": group_val_str, "is_estimate": is_estimate}

    trim = spec.get("trim_trailing_blank", False)
    records = []
    for row in range(data_start_row, ws.max_row + 1):
        id_value = ws.cell(row=row, column=id_idx).value
        if trim and id_value is None:
            continue
        for col, plan in col_plan.items():
            value = ws.cell(row=row, column=col).value
            records.append(
                {
                    "_row": row,
                    id_name: id_value,
                    group_role: plan["group"],
                    "metric": plan["name"],
                    "value": value,
                    "is_estimate": plan["is_estimate"],
                    "_role": plan["role"],
                }
            )
    return pd.DataFrame(records)


def parse_sparse_tracker(ws, spec: dict) -> pd.DataFrame:
    header_row = spec.get("header_row", 1)
    data_start_row = spec.get("data_start_row", header_row + 1)
    max_col = ws.max_column
    headers = unmerge_row_values(ws, header_row, max_col)
    header_index = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h is not None}

    year_idx = header_index[spec["date_construction"]["year_column"].strip().lower()]
    month_idx = header_index[spec["date_construction"]["month_column"].strip().lower()]
    col_specs = {k.strip().lower(): v for k, v in spec.get("columns", {}).items()}

    records = []
    for row in range(data_start_row, ws.max_row + 1):
        year_val = ws.cell(row=row, column=year_idx).value
        month_val = ws.cell(row=row, column=month_idx).value
        if year_val is None and month_val is None:
            continue
        record = {
            "_row": row,
            "year": year_val,
            "month": month_val,
            "date": _safe_date(year_val, _month_to_num(month_val)) if year_val and month_val else pd.NaT,
        }
        for col in range(1, max_col + 1):
            if col in (year_idx, month_idx):
                continue
            header_text = headers[col - 1]
            if header_text is None:
                continue
            info = col_specs.get(str(header_text).strip().lower())
            if info is None:
                continue
            record[info["rename"]] = ws.cell(row=row, column=col).value
        records.append(record)
    return pd.DataFrame(records)


LAYOUT_PARSERS = {
    "long": parse_long,
    "wide_pivot_year_month": parse_wide_pivot_year_month,
    "wide_pivot_year_cols": parse_wide_pivot_year_cols,
    "wide_pivot_merged_header": parse_wide_pivot_merged_header,
    "sparse_tracker": parse_sparse_tracker,
}


def parse_workbook(wb_values, sheet_config: dict) -> tuple[dict, dict]:
    """Parse every configured sheet found in the workbook.

    Returns (parsed: {sheet_name: DataFrame}, skipped: {sheet_name: reason}).
    """
    parsed = {}
    skipped = {}
    for sheet_name, spec in sheet_config.get("sheets", {}).items():
        if sheet_name not in wb_values.sheetnames:
            skipped[sheet_name] = "configured but not found in workbook"
            logger.warning("Configured sheet '%s' not found in workbook", sheet_name)
            continue

        layout = spec["layout"]
        if layout == "derived_skip":
            skipped[sheet_name] = spec.get("reason", "derived/computed sheet, not ingested as raw source")
            logger.info("Skipping derived sheet '%s'", sheet_name)
            continue

        parser = LAYOUT_PARSERS.get(layout)
        if parser is None:
            skipped[sheet_name] = f"unknown layout '{layout}'"
            logger.error("Unknown layout '%s' for sheet '%s'", layout, sheet_name)
            continue

        try:
            df = parser(wb_values[sheet_name], spec)
            parsed[sheet_name] = df
            logger.info("Parsed sheet '%s': %d rows", sheet_name, len(df))
        except Exception as exc:  # noqa: BLE001 - one bad sheet must not crash the whole refresh
            skipped[sheet_name] = f"parse error: {exc}"
            logger.error("Failed to parse sheet '%s': %s", sheet_name, exc, exc_info=True)

    return parsed, skipped
