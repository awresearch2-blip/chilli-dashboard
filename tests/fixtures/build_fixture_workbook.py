"""Builds a small synthetic .xlsx replicating the real workbook's structural
quirks, for tests only. None of these values are real market data.
"""

import datetime as dt

import openpyxl


def build_fixture_workbook(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- LongSheet: simple date-indexed table with an error token, a
    # negative value (bad data on purpose), and trailing blank rows.
    ws = wb.create_sheet("LongSheet")
    ws.append(["Date", "Price", "Arrivals"])
    ws.append([dt.datetime(2024, 1, 1), 100, 500])
    ws.append([dt.datetime(2024, 1, 2), "Closed", 510])
    ws.append([dt.datetime(2024, 1, 3), 105, -20])  # negative arrival: bad data
    ws.append([dt.datetime(2024, 1, 4), "110", 520])  # number stored as text
    ws.append([dt.datetime(2024, 1, 5), "TBD", 530])  # unrecognized text -- must NOT be guessed at
    ws.append([dt.datetime(2024, 1, 1), 100, 500])  # exact duplicate of the first row
    ws.append([None, None, None])
    ws.append([None, None, None])

    # --- MergedSheet: 2-row merged header (2 varieties x Low/High), plus a
    # formula-error token cell.
    ws = wb.create_sheet("MergedSheet")
    ws.append(["Date", "VarietyA", None, "VarietyB", None])
    ws.merge_cells("B1:C1")
    ws.merge_cells("D1:E1")
    ws.append([None, "Low", "High", "Low", "High"])
    ws.append([dt.datetime(2024, 1, 1), 10, 20, 30, 40])
    ws.append([dt.datetime(2024, 1, 2), "#DIV/0!", 22, 31, 41])

    # --- PivotSheet: Year rows x Month columns, with a non-year legend row
    # that must be excluded by the id_is_4digit_year filter.
    ws = wb.create_sheet("PivotSheet")
    ws.append(["Year", "Jan", "Feb"])
    ws.append([2024, 100, 110])
    ws.append([2025, 105, 115])
    ws.append(["legend: low = bad", None, None])

    # --- StackedTablesSheet: two independent Month x Year tables separated
    # by a blank row -- parsing must stop at the first table.
    ws = wb.create_sheet("StackedTablesSheet")
    ws.append(["Months", 2024, 2025])
    ws.append(["Jan", 1000, 1100])
    ws.append(["Feb", 900, 950])
    ws.append([None, None, None])
    ws.append(["Months (converted unit)", 2024, 2025])
    ws.append(["Jan", 450, 495])

    # --- UnconfiguredSheet: present in the workbook but absent from the
    # test's sheet config, to exercise unconfigured-sheet detection.
    ws = wb.create_sheet("UnconfiguredSheet")
    ws.append(["A", "B"])
    ws.append([1, 2])

    wb.save(path)
